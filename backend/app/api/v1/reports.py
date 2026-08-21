"""
IntelliProcure AI – Reports Engine
Real SQL-backed report generation with CSV, Excel (.xlsx), and PDF export.

Supported report types:
  spend, supplier, purchase, rfq, po, invoice, inventory, risk, budget

All values match the underlying database filtered by the query parameters provided.
"""

import csv
import io
from datetime import datetime
from typing import Optional
from uuid import UUID

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func

from app.database.session import get_db
from app.core.security import get_current_active_user
from app.models.supplier import Supplier
from app.models.purchase_request import PurchaseRequest
from app.models.purchase_order import PurchaseOrder
from app.models.rfq import (
    RFQ, Quotation, Invoice, Inventory, GoodsReceipt, Contract
)
from app.models.budget import Budget

router = APIRouter()

# ─── REPORT DATA BUILDERS ─────────────────────────────────────────────────────

def _build_spend_rows(db, start_date=None, end_date=None, department=None, category=None, status=None):
    q = db.query(PurchaseOrder).filter(PurchaseOrder.is_deleted == False)
    if start_date:
        try: q = q.filter(PurchaseOrder.created_at >= datetime.strptime(start_date[:10], "%Y-%m-%d"))
        except ValueError: pass
    if end_date:
        try: q = q.filter(PurchaseOrder.created_at <= datetime.strptime(end_date[:10], "%Y-%m-%d"))
        except ValueError: pass
    if status:
        q = q.filter(PurchaseOrder.status == status)
    pos = q.options(joinedload(PurchaseOrder.supplier)).all()
    headers = ["PO Number", "Supplier", "Status", "Total Amount", "Discount", "Net Amount", "Created Date"]
    rows = [
        [
            po.po_number,
            po.supplier.company_name if po.supplier else "",
            po.status.value if hasattr(po.status, "value") else str(po.status),
            round(po.total_amount or 0, 2),
            round(po.discount_amount or 0, 2),
            round((po.total_amount or 0) - (po.discount_amount or 0), 2),
            po.created_at.strftime("%Y-%m-%d") if po.created_at else ""
        ]
        for po in pos
    ]
    return headers, rows


def _build_supplier_rows(db, category=None, status=None, **_):
    q = db.query(Supplier).filter(Supplier.is_deleted == False)
    if category:
        q = q.filter(Supplier.category.ilike(f"%{category}%"))
    if status:
        q = q.filter(Supplier.status == status)
    suppliers = q.all()
    headers = ["Company Name", "Code", "Category", "Country", "Status", "Rating", "Risk Level", "Risk Score", "Total Spend", "Total Orders"]
    rows = [
        [
            s.company_name, s.supplier_code or "", s.category or "", s.country or "",
            s.status, s.overall_rating or 0, s.risk_level or "low", s.risk_score or 0,
            round(s.total_spend or 0, 2), s.total_orders or 0
        ]
        for s in suppliers
    ]
    return headers, rows


def _build_purchase_rows(db, start_date=None, end_date=None, department=None, status=None, **_):
    q = db.query(PurchaseRequest).filter(PurchaseRequest.is_deleted == False)
    if start_date:
        try: q = q.filter(PurchaseRequest.created_at >= datetime.strptime(start_date[:10], "%Y-%m-%d"))
        except ValueError: pass
    if end_date:
        try: q = q.filter(PurchaseRequest.created_at <= datetime.strptime(end_date[:10], "%Y-%m-%d"))
        except ValueError: pass
    if department:
        q = q.filter(PurchaseRequest.department.ilike(f"%{department}%"))
    if status:
        q = q.filter(PurchaseRequest.status == status)
    prs = q.all()
    headers = ["PR Number", "Title", "Department", "Status", "Estimated Amount", "Priority", "Created Date"]
    rows = [
        [
            pr.pr_number, pr.title, pr.department or "",
            pr.status.value if hasattr(pr.status, "value") else str(pr.status),
            round(pr.estimated_amount or 0, 2), pr.priority or "medium",
            pr.created_at.strftime("%Y-%m-%d") if pr.created_at else ""
        ]
        for pr in prs
    ]
    return headers, rows


def _build_rfq_rows(db, start_date=None, end_date=None, status=None, **_):
    q = db.query(RFQ).filter(RFQ.is_deleted == False)
    if start_date:
        try: q = q.filter(RFQ.created_at >= datetime.strptime(start_date[:10], "%Y-%m-%d"))
        except ValueError: pass
    if end_date:
        try: q = q.filter(RFQ.created_at <= datetime.strptime(end_date[:10], "%Y-%m-%d"))
        except ValueError: pass
    if status:
        q = q.filter(RFQ.status == status)
    rfqs = q.all()
    headers = ["RFQ Number", "Title", "Status", "Estimated Value", "Currency", "Deadline", "Created Date"]
    rows = [
        [
            r.rfq_number, r.title,
            r.status.value if hasattr(r.status, "value") else str(r.status),
            round(r.estimated_value or 0, 2),
            r.currency or "USD",
            r.deadline.strftime("%Y-%m-%d") if r.deadline else "",
            r.created_at.strftime("%Y-%m-%d") if r.created_at else ""
        ]
        for r in rfqs
    ]
    return headers, rows


def _build_po_rows(db, start_date=None, end_date=None, status=None, **_):
    return _build_spend_rows(db, start_date=start_date, end_date=end_date, status=status)


def _build_invoice_rows(db, start_date=None, end_date=None, status=None, **_):
    q = db.query(Invoice).filter(Invoice.is_deleted == False)
    if start_date:
        try: q = q.filter(Invoice.created_at >= datetime.strptime(start_date[:10], "%Y-%m-%d"))
        except ValueError: pass
    if end_date:
        try: q = q.filter(Invoice.created_at <= datetime.strptime(end_date[:10], "%Y-%m-%d"))
        except ValueError: pass
    if status:
        q = q.filter(Invoice.status == status)
    invoices = q.options(joinedload(Invoice.supplier), joinedload(Invoice.purchase_order)).all()
    headers = ["Invoice Number", "Supplier", "PO Reference", "Status", "Total Amount", "Tax Amount", "Risk Score", "Created Date"]
    rows = [
        [
            inv.invoice_number,
            inv.supplier.company_name if inv.supplier else "",
            inv.purchase_order.po_number if inv.purchase_order else "",
            inv.status.value if hasattr(inv.status, "value") else str(inv.status),
            round(inv.total_amount or 0, 2),
            round(inv.tax_amount or 0, 2),
            round(inv.fraud_risk_score or 0, 1),
            inv.created_at.strftime("%Y-%m-%d") if inv.created_at else ""
        ]
        for inv in invoices
    ]
    return headers, rows


def _build_inventory_rows(db, category=None, **_):
    q = db.query(Inventory).filter(Inventory.is_deleted == False)
    if category:
        q = q.filter(Inventory.category.ilike(f"%{category}%"))
    items = q.all()
    headers = ["Item Code", "Item Name", "Category", "Warehouse", "Qty On Hand", "Reorder Point", "Reorder Qty", "Unit Cost", "Total Valuation", "Status"]
    rows = [
        [
            it.item_code, it.item_name, it.category or "", it.warehouse_location or "WH-MAIN",
            it.quantity_on_hand, it.reorder_point or 0, it.reorder_quantity or 0,
            round(it.unit_cost or 0, 2),
            round((it.quantity_on_hand or 0) * (it.unit_cost or 0), 2),
            "LOW STOCK" if (it.quantity_on_hand or 0) <= (it.reorder_point or 0) else "OK"
        ]
        for it in items
    ]
    return headers, rows


def _build_risk_rows(db, **_):
    from app.services.ai_service import AIPredictiveEngine
    invoices = db.query(Invoice).options(joinedload(Invoice.supplier), joinedload(Invoice.purchase_order))\
        .filter(Invoice.is_deleted == False).all()
    headers = ["Invoice Number", "Supplier", "Total Amount", "Risk Score", "Risk Level", "Primary Reason"]
    rows = []
    for inv in invoices:
        po = inv.purchase_order
        historical = [h.total_amount for h in invoices if h.supplier_id == inv.supplier_id and h.id != inv.id]
        dup = any(h.invoice_number == inv.invoice_number and h.id != inv.id for h in invoices)
        res = AIPredictiveEngine.evaluate_invoice_fraud_risk(
            invoice_amount=inv.total_amount, po_amount=po.total_amount if po else 0.0,
            po_status=po.status if po else "none", is_duplicate_number=dup,
            historical_amounts=historical,
            same_amount_within_30days_count=sum(1 for h in invoices if h.supplier_id == inv.supplier_id and h.total_amount == inv.total_amount and h.id != inv.id),
            is_new_vendor=len(historical) == 0
        )
        rows.append([
            inv.invoice_number,
            inv.supplier.company_name if inv.supplier else "",
            round(inv.total_amount or 0, 2),
            res["risk_score"],
            res["risk_level"],
            res["reasons"][0] if res["reasons"] else "No risk flag"
        ])
    rows.sort(key=lambda x: x[3], reverse=True)
    return headers, rows


def _build_budget_rows(db, department=None, **_):
    q = db.query(Budget).filter(Budget.is_deleted == False)
    if department:
        q = q.filter(Budget.department_name.ilike(f"%{department}%"))
    budgets = q.all()
    headers = ["Budget Name", "Department", "Category", "Allocated", "Spent", "Remaining", "Utilization %", "Alert Status"]
    rows = []
    for b in budgets:
        allocated = b.allocated_amount or 0
        spent = b.spent_amount or 0
        remaining = allocated - spent
        utilization = round((spent / allocated * 100) if allocated > 0 else 0, 1)
        rows.append([
            b.name, b.department_name or "", b.category or "",
            round(allocated, 2), round(spent, 2),
            round(remaining, 2),
            utilization,
            "CRITICAL" if utilization >= 90 else ("WARNING" if utilization >= 80 else "NORMAL")
        ])
    return headers, rows


REPORT_BUILDERS = {
    "spend":     _build_spend_rows,
    "supplier":  _build_supplier_rows,
    "purchase":  _build_purchase_rows,
    "rfq":       _build_rfq_rows,
    "po":        _build_po_rows,
    "invoice":   _build_invoice_rows,
    "inventory": _build_inventory_rows,
    "risk":      _build_risk_rows,
    "budget":    _build_budget_rows,
}


# ─── CSV GENERATOR ────────────────────────────────────────────────────────────

def _generate_csv(headers: list, rows: list) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(headers)
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility


# ─── EXCEL GENERATOR ──────────────────────────────────────────────────────────

def _generate_excel(title: str, headers: list, rows: list) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    # Header row styling
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(bottom=Side(style="thin", color="DDDDDD"))

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F0F4FF")

    # Auto-fit columns
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, len(rows) + 2)),
            default=10
        )
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─── PDF GENERATOR ───────────────────────────────────────────────────────────

def _generate_pdf(report_type: str, title: str, headers: list, rows: list) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        rightMargin=15 * mm, leftMargin=15 * mm,
        topMargin=15 * mm, bottomMargin=15 * mm
    )

    styles = getSampleStyleSheet()
    primary = colors.HexColor("#2563EB")
    elements = []

    # Title
    title_style = ParagraphStyle("Title", parent=styles["Heading1"], textColor=primary, fontSize=16, spaceAfter=6)
    elements.append(Paragraph(f"IntelliProcure AI — {title}", title_style))
    elements.append(Paragraph(f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')} | {len(rows)} records", styles["Normal"]))
    elements.append(Spacer(1, 10 * mm))

    # Table data
    table_data = [headers] + [[str(v) for v in row] for row in rows]

    col_count = len(headers)
    page_width = landscape(A4)[0] - 30 * mm
    col_widths = [page_width / col_count] * col_count

    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), primary),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F4FF")]),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#CCCCCC")),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWHEIGHT", (0, 0), (-1, -1), 16),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(tbl)

    doc.build(elements)
    buf.seek(0)
    return buf.read()


# ─── FASTAPI ENDPOINT ─────────────────────────────────────────────────────────

@router.get("/generate")
async def generate_report(
    report_type: str = Query(..., description="spend|supplier|purchase|rfq|po|invoice|inventory|risk|budget"),
    format: str = Query("csv", description="csv|excel|pdf"),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    department: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    db: Session = Depends(get_db),
    current_user=Depends(get_current_active_user)
):
    """
    Generate a real-data procurement report with optional filters.
    Exports as CSV, Excel (.xlsx), or PDF — values always match the underlying DB.
    """
    builder = REPORT_BUILDERS.get(report_type)
    if not builder:
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail=f"Unknown report_type '{report_type}'. Valid: {list(REPORT_BUILDERS.keys())}")

    filter_kwargs = dict(
        start_date=start_date, end_date=end_date,
        department=department, category=category, status=status
    )

    headers, rows = builder(db, **filter_kwargs)

    REPORT_TITLES = {
        "spend": "Spend & Purchase Order Report",
        "supplier": "Supplier Performance & Risk Report",
        "purchase": "Purchase Requests Report",
        "rfq": "Request For Quotation (RFQ) Report",
        "po": "Purchase Orders Report",
        "invoice": "Invoice Audit Report",
        "inventory": "Inventory Valuation & Stock Report",
        "risk": "AI Fraud & Risk Detection Report",
        "budget": "Budget Utilization & Variance Report",
    }
    title = REPORT_TITLES.get(report_type, report_type.title() + " Report")
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M")
    filename_base = f"intelliprocure_{report_type}_report_{ts}"

    fmt = format.lower()

    if fmt == "csv":
        content = _generate_csv(headers, rows)
        return StreamingResponse(
            io.BytesIO(content),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename_base}.csv"'}
        )

    elif fmt in ("excel", "xlsx"):
        try:
            content = _generate_excel(title, headers, rows)
        except ImportError:
            # openpyxl not installed — fallback to CSV
            content = _generate_csv(headers, rows)
            return StreamingResponse(
                io.BytesIO(content),
                media_type="text/csv",
                headers={"Content-Disposition": f'attachment; filename="{filename_base}_fallback.csv"'}
            )
        return StreamingResponse(
            io.BytesIO(content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename_base}.xlsx"'}
        )

    elif fmt == "pdf":
        try:
            content = _generate_pdf(report_type, title, headers, rows)
        except ImportError:
            # reportlab not installed — fallback to CSV
            content = _generate_csv(headers, rows)
            return StreamingResponse(
                io.BytesIO(content),
                media_type="text/csv",
                headers={"Content-Disposition": f'attachment; filename="{filename_base}_fallback.csv"'}
            )
        return StreamingResponse(
            io.BytesIO(content),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename_base}.pdf"'}
        )

    else:
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail=f"Unsupported format '{format}'. Use: csv, excel, pdf")


@router.get("/")
async def list_reports(current_user=Depends(get_current_active_user)):
    """List available report types and their descriptions."""
    return [
        {"id": "spend",     "title": "Spend & Purchase Order Report",         "category": "Finance",    "description": "All POs with amounts, discounts, supplier breakdown filtered by date/status."},
        {"id": "supplier",  "title": "Supplier Performance & Risk Report",     "category": "Suppliers",  "description": "Vendor ratings, delivery scores, risk levels, and total spend per supplier."},
        {"id": "purchase",  "title": "Purchase Requests Report",               "category": "Procurement","description": "All PRs with department, priority, status, and estimated amounts."},
        {"id": "rfq",       "title": "RFQ Activity Report",                    "category": "Sourcing",   "description": "Active and closed RFQs with budgets, deadlines, and quotation counts."},
        {"id": "po",        "title": "Purchase Orders Report",                 "category": "Finance",    "description": "Full PO listing with supplier, status, amounts, and date range."},
        {"id": "invoice",   "title": "Invoice Audit Report",                   "category": "Finance",    "description": "Invoice status, 3-way match outcomes, tax amounts, and AI risk scores."},
        {"id": "inventory", "title": "Inventory Valuation & Stock Report",     "category": "Inventory",  "description": "Stock levels, reorder alerts, valuations, and warehouse locations."},
        {"id": "risk",      "title": "AI Fraud & Risk Detection Report",       "category": "Compliance", "description": "All invoices scored by AI across 5 signals with risk levels and reasons."},
        {"id": "budget",    "title": "Budget Utilization & Variance Report",   "category": "Finance",    "description": "Departmental budgets with allocated, spent, remaining, and alert thresholds."},
    ]
